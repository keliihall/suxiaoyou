from __future__ import annotations

from dataclasses import replace
import hashlib
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.office_rendering import (
    OfficeRenderCache,
    ProviderAvailability,
    RenderManifest,
    RendererDescriptor,
    RenderRequest,
)
from app.office_validation import (
    OfficeDraftStaleError,
    OfficeDraftValidationService,
    OfficeGoldenPolicy,
    VisualDiffPolicy,
)
from app.schemas.agent import AgentInfo
from app.tool import workspace_transaction as transaction_module
from app.tool.context import ToolContext
from app.tool.workspace_transaction import WorkspaceMutationTransaction
from tests.test_office_rendering.helpers import png_bytes, write_render_artifacts
from tests.test_office_templates.helpers import (
    make_docx_template,
    manifest_for,
    rewrite_zip,
    zip_entries,
)


pytestmark = [
    pytest.mark.asyncio,
    pytest.mark.skipif(
        transaction_module.guarded_file_mutation_unavailable_reason() is not None,
        reason="guarded mutation primitive unavailable",
    ),
]


class DraftProvider:
    def __init__(
        self,
        *,
        quality: str = "authoritative",
        colors: dict[str, int] | None = None,
    ) -> None:
        self.colors = colors or {}
        self._descriptor = RendererDescriptor(
            renderer_id="draft-test-renderer",
            renderer_version="1",
            font_digest="f" * 64,
            quality=quality,  # type: ignore[arg-type]
        )

    @property
    def descriptor(self) -> RendererDescriptor:
        return self._descriptor

    def availability(self) -> ProviderAvailability:
        return ProviderAvailability(available=True)

    async def render(self, request: RenderRequest, output_dir: Path) -> RenderManifest:
        content = png_bytes(red=self.colors.get(request.source_sha256, 24))
        pdf, pages = write_render_artifacts(output_dir, (content,))
        return RenderManifest.for_request(
            request,
            self._descriptor,
            pages,
            pdf=pdf,
        )


def _context(workspace: Path) -> ToolContext:
    return ToolContext(
        session_id="session",
        message_id="message",
        agent=AgentInfo(name="test", description="", mode="primary"),
        call_id="office-call",
        workspace=str(workspace),
    )


def _variant(content: bytes, old: bytes, new: bytes) -> bytes:
    entries = zip_entries(content)
    document = entries["word/document.xml"].replace(old, new, 1)
    assert document != entries["word/document.xml"]
    return rewrite_zip(
        content,
        replacements={"word/document.xml": document},
    )


def _service(tmp_path: Path, provider: DraftProvider) -> OfficeDraftValidationService:
    return OfficeDraftValidationService(
        cache=OfficeRenderCache((tmp_path / "cache").resolve()),
        provider=provider,
        parameters_version="draft-validation-v1",
        parameters={"dpi": 144},
    )


async def test_staged_edit_passes_authoritative_gates_without_touching_original(
    tmp_path: Path,
) -> None:
    workspace = (tmp_path / "workspace").resolve()
    private = (tmp_path / "private").resolve()
    workspace.mkdir()
    target = workspace / "report.docx"
    baseline_bytes = make_docx_template()
    candidate_bytes = _variant(
        baseline_bytes,
        "正文".encode(),
        "报告".encode(),
    )
    target.write_bytes(baseline_bytes)
    transaction = WorkspaceMutationTransaction(
        workspace,
        _context(workspace),
        operation="office.edit",
        storage_root=private,
    )
    staged = transaction.prepare_paths([target])
    staged_target = transaction.staged_path(target)
    view = transaction.arm_office_precommit_validation(target)
    staged_target.write_bytes(candidate_bytes)
    service = _service(tmp_path, DraftProvider())

    baseline = await service.capture(
        boundary_root=workspace,
        source_path=target,
    )
    candidate = await service.capture(
        boundary_root=staged,
        source_path=staged_target,
    )
    result = service.compare(
        baseline=baseline,
        candidate=candidate,
        allowed_changed_parts=("word/document.xml",),
        visual_policy=VisualDiffPolicy(max_blank_fraction_increase=1.0),
    )
    report = result.report

    assert report.verdict == "pass"
    assert target.read_bytes() == baseline_bytes
    assert transaction.collect_changes().writes == ("report.docx",)
    assert result.commit_seal is None
    assert result.candidate.source_sha256 == hashlib.sha256(
        candidate_bytes
    ).hexdigest()
    commit = transaction.commit_with_precommit_office_seal(
        replace(
            result.candidate,
            validation_generation=view.validation_generation,
        )
    )
    assert commit.written_files == (str(target),)
    assert target.read_bytes() == candidate_bytes


async def test_staged_create_requires_explicit_golden_and_remains_private(
    tmp_path: Path,
) -> None:
    workspace = (tmp_path / "workspace").resolve()
    golden_root = (tmp_path / "golden").resolve()
    private = (tmp_path / "private").resolve()
    workspace.mkdir()
    golden_root.mkdir()
    target = workspace / "suxiaoyou_written" / "report.docx"
    golden_path = golden_root / "template.docx"
    golden_bytes = make_docx_template()
    candidate_bytes = _variant(
        golden_bytes,
        "正文".encode(),
        "报告".encode(),
    )
    golden_path.write_bytes(golden_bytes)
    transaction = WorkspaceMutationTransaction(
        workspace,
        _context(workspace),
        operation="office.create",
        storage_root=private,
    )
    staged = transaction.prepare_paths([target])
    staged_target = transaction.staged_path(target)
    view = transaction.arm_office_precommit_validation(target)
    staged_target.parent.mkdir()
    staged_target.write_bytes(candidate_bytes)
    service = _service(tmp_path, DraftProvider())

    golden = await service.capture(
        boundary_root=golden_root,
        source_path=golden_path,
    )
    candidate = await service.capture(
        boundary_root=staged,
        source_path=staged_target,
    )
    template_manifest = manifest_for(
        golden_bytes,
        "docx",
        ("body", "client", "footer", "header", "table"),
        template_id="report",
        version="1.0.0",
    )
    policy = OfficeGoldenPolicy(
        policy_id="first-party/report/1",
        template_id="report",
        template_version="1.0.0",
        template_manifest_sha256=template_manifest.template_sha256,
        baseline_sha256=golden.source_sha256,
        renderer_id=golden.manifest.renderer_id,
        renderer_version=golden.manifest.renderer_version,
        font_digest=golden.manifest.font_digest,
        parameters_version=golden.manifest.parameters_version,
        parameters_sha256=golden.manifest.parameters_sha256,
        allowed_changed_parts=("word/document.xml",),
        visual=VisualDiffPolicy(max_blank_fraction_increase=1.0),
    )
    result = service.compare_with_golden(
        golden=golden,
        candidate=candidate,
        policy=policy,
        template_manifest=template_manifest,
    )
    report = result.report

    assert report.verdict == "pass"
    assert not target.exists()
    with pytest.raises(OfficeDraftStaleError, match="golden"):
        service.compare_with_golden(
            golden=golden,
            candidate=candidate,
            policy=replace(policy, renderer_version="different"),
            template_manifest=template_manifest,
        )
    assert result.commit_seal is None
    transaction.commit_with_precommit_office_seal(
        replace(
            result.candidate,
            validation_generation=view.validation_generation,
        )
    )
    assert target.read_bytes() == candidate_bytes


async def test_logical_sheet_delta_rejects_unrequested_sheet_deletion(
    tmp_path: Path,
) -> None:
    baseline_root = (tmp_path / "baseline-xlsx").resolve()
    candidate_root = (tmp_path / "candidate-xlsx").resolve()
    baseline_root.mkdir()
    candidate_root.mkdir()
    baseline_path = baseline_root / "book.xlsx"
    candidate_path = candidate_root / "book.xlsx"

    baseline_book = Workbook()
    baseline_book.active.title = "Keep"
    baseline_book.create_sheet("Tail")
    baseline_book.save(baseline_path)
    baseline_book.close()
    candidate_book = Workbook()
    candidate_book.active.title = "Keep"
    candidate_book.save(candidate_path)
    candidate_book.close()

    service = _service(tmp_path, DraftProvider())
    baseline = await service.capture(
        boundary_root=baseline_root,
        source_path=baseline_path,
    )
    candidate = await service.capture(
        boundary_root=candidate_root,
        source_path=candidate_path,
    )
    result = service.compare(
        baseline=baseline,
        candidate=candidate,
        allowed_changed_parts=("*",),
        visual_policy=VisualDiffPolicy(max_blank_fraction_increase=1.0),
        expected_logical_unit_delta=0,
    )

    logical = next(
        check for check in result.report.checks if check.code == "logical_unit_delta"
    )
    assert logical.outcome == "fail"
    assert result.report.verdict == "fail"
    assert result.commit_seal is None


async def test_approximate_draft_never_passes_and_stale_capture_is_rejected(
    tmp_path: Path,
) -> None:
    baseline_root = (tmp_path / "baseline").resolve()
    candidate_root = (tmp_path / "candidate").resolve()
    baseline_root.mkdir()
    candidate_root.mkdir()
    baseline_path = baseline_root / "report.docx"
    candidate_path = candidate_root / "report.docx"
    baseline_bytes = make_docx_template()
    candidate_bytes = _variant(
        baseline_bytes,
        "正文".encode(),
        "报告".encode(),
    )
    later_bytes = _variant(
        candidate_bytes,
        "报告".encode(),
        "异常".encode(),
    )
    baseline_path.write_bytes(baseline_bytes)
    candidate_path.write_bytes(candidate_bytes)
    service = _service(tmp_path, DraftProvider(quality="approximate"))
    baseline = await service.capture(
        boundary_root=baseline_root,
        source_path=baseline_path,
    )
    candidate = await service.capture(
        boundary_root=candidate_root,
        source_path=candidate_path,
    )

    result = service.compare(
        baseline=baseline,
        candidate=candidate,
        allowed_changed_parts=("word/document.xml",),
        visual_policy=VisualDiffPolicy(max_blank_fraction_increase=1.0),
    )
    report = result.report
    assert report.verdict == "needs_review"
    assert result.commit_seal is None
    assert next(
        item for item in report.checks if item.code == "authoritative_quality"
    ).outcome == "needs_review"

    candidate_path.write_bytes(later_bytes)
    with pytest.raises(OfficeDraftStaleError, match="changed"):
        service.compare(
            baseline=baseline,
            candidate=candidate,
            allowed_changed_parts=("word/document.xml",),
            visual_policy=VisualDiffPolicy(max_blank_fraction_increase=1.0),
        )


async def test_seeded_visual_draft_defect_returns_evidence_box(
    tmp_path: Path,
) -> None:
    workspace = (tmp_path / "workspace").resolve()
    private = (tmp_path / "private").resolve()
    workspace.mkdir()
    baseline_path = workspace / "report.docx"
    baseline_bytes = make_docx_template()
    candidate_bytes = _variant(
        baseline_bytes,
        "正文".encode(),
        "报告".encode(),
    )
    baseline_path.write_bytes(baseline_bytes)
    transaction = WorkspaceMutationTransaction(
        workspace,
        _context(workspace),
        operation="office.edit",
        storage_root=private,
    )
    staged = transaction.prepare_paths([baseline_path])
    candidate_path = transaction.staged_path(baseline_path)
    candidate_path.write_bytes(candidate_bytes)
    colors = {
        hashlib.sha256(baseline_bytes).hexdigest(): 10,
        hashlib.sha256(candidate_bytes).hexdigest(): 240,
    }
    service = _service(tmp_path, DraftProvider(colors=colors))
    baseline = await service.capture(
        boundary_root=workspace,
        source_path=baseline_path,
    )
    candidate = await service.capture(
        boundary_root=staged,
        source_path=candidate_path,
    )

    result = service.compare(
        baseline=baseline,
        candidate=candidate,
        allowed_changed_parts=("word/document.xml",),
        visual_policy=VisualDiffPolicy(max_blank_fraction_increase=1.0),
    )
    report = result.report

    pixel = next(item for item in report.checks if item.code == "pixel_delta")
    assert report.verdict == "fail"
    assert result.commit_seal is None
    assert pixel.box is not None
    assert pixel.box.to_dict() == {
        "page_number": 1,
        "x": 0,
        "y": 0,
        "width": 2,
        "height": 2,
    }
    transaction.abort()
    assert baseline_path.read_bytes() == baseline_bytes
